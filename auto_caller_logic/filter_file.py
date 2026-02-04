"""
Filter file generator for Google Drive uploads.

This module contains the FilterFile class that generates Excel workbooks
with Hebrew headers and ARRAYFORMULA formulas.
"""

import io
import sys
import os
from typing import Dict, Any, Optional, List
from pathlib import Path
from .google_drive_utils import BaseProcess
from .config import _get_default_config
from common_utils.config_manager import ConfigManager
from .spreadsheet_updaters.base import BaseSpreadsheetUpdater
from .spreadsheet_updaters.gap_spreadsheet_updater import GapSpreadsheetUpdater
from .mail_service import create_mail_service
from common_utils.db_connection import DatabaseConnection
class FilterFile(BaseProcess):

    def __init__(self, drive_service, config_manager: ConfigManager, name: str, spreadsheet_updaters: List[BaseSpreadsheetUpdater], mail_service, customers_google_folder_id: str, customers_file_name_pattern: str, auto_dialer_file_name_pattern: str, allowed_gaps_sheet_config: Dict[str, Any]):
        super().__init__(drive_service, config_manager, name, spreadsheet_updaters=spreadsheet_updaters, mail_service=mail_service)
        
        self.customers_google_folder_id = customers_google_folder_id
        self.customers_file_name_pattern = customers_file_name_pattern
        self.auto_dialer_file_name_pattern = auto_dialer_file_name_pattern
        self.allowed_gaps_sheet_config = allowed_gaps_sheet_config
        # Store summarize_data for gaps sheet insertion
        self._summarize_data = None

    def generate_data(self, **kwargs):
        """
        Generate an Excel workbook with Hebrew headers and ARRAYFORMULA formulas.

        Returns:
            BytesIO buffer containing the Excel file
        """
        try:
            calls = kwargs.get('calls')
            caller_id = kwargs.get('caller_id')
            input_file_path = kwargs.get('customers_input_file')
            nick_name = kwargs.get('nick_name')
            customers, customers_input_file = self._get_customers(input_file_path)

            print(f"nickname: {nick_name}", file=sys.stderr)

            print(f"Input file path: {customers_input_file} caller id: {caller_id}", file=sys.stderr)
        
            # Create summarize_data with header values (A1-A4): [date, time, customers_input_file, caller_id]
            from datetime import datetime
            current_datetime = datetime.now()
            date_str = current_datetime.strftime("%d.%m.%Y")
            time_str = current_datetime.strftime("%H:%M")
            summarize_data = {
                'date_str': date_str,
                'time_str': time_str,
                'customers_input_file_name': customers_input_file,
                'caller_id': caller_id,
                'nick_name': nick_name
            }
            
            # Store for later use in gaps sheet insertion
            self._summarize_data = summarize_data
        
            return {
                'calls': calls,
                'customers': customers,
                'summarize_data': summarize_data
            }

        except ImportError:
            print("Error: openpyxl is required. Install it with: pip install openpyxl", file=sys.stderr)
            raise ImportError("openpyxl is required. Install it with: pip install openpyxl")    
        except Exception as e:
            print(f"Error creating Auto Calls Excel workbook: {e}", file=sys.stderr)
            raise RuntimeError(f"Error creating Auto Calls Excel workbook: {e}")

    def post_process_implementation(self, excel_info: Dict[str, Any]):
        """The specific logic to post process the file."""
        
        # Safely access nested dictionary
        filter_info = excel_info.get('filter')
        if filter_info is None:
            raise ValueError("Filter workbook info is not found in excel_info")
        
        filter_google_sheet_id = filter_info.get('file_id')
        filter_google_sheet_first_sheet_id = filter_info.get('sheet_id')
        print(f"Filter Google sheet ID: {filter_google_sheet_id}", file=sys.stderr)

        if filter_google_sheet_id is None:
            raise ValueError("Filter Google sheet ID is not found")

        callers_gap = self.get_list_of_missing_customers(filter_google_sheet_id, filter_google_sheet_first_sheet_id)

        print (f"finished .... post process implementation. len of callers gap: {len(callers_gap)}", file=sys.stderr)
        
        allowed_gaps = self._get_allowed_gaps_list()

        print(f"allowed_gaps length: {len(allowed_gaps)}", file=sys.stderr)

        filtered_callers_gap = [
            item for item in callers_gap 
            if str(item).strip() not in allowed_gaps
        ]

        print(f"filtered_callers_gap: {filtered_callers_gap}", file=sys.stderr)

        return {'callers_gap': filtered_callers_gap, "metadata": self._summarize_data}

    def get_mail_data(self) -> Dict[str, Any]:
        if not self.post_data:
            return {}
        
        callers_gap = self.post_data.get('callers_gap')
        
        return {
            'number_of_gaps': len(callers_gap),
            'nick_name': self._summarize_data.get('nick_name', '')
        }

    def get_generated_data(self) -> Dict[str, Any]:
        return self._summarize_data
    
    def _get_sheet_name_from_id(self, spreadsheet_id: str, sheet_id: int) -> str:
        """
        Get the sheet name for a given sheet ID.
        
        Args:
            spreadsheet_id: Google Sheets spreadsheet ID
            sheet_id: Sheet ID (integer)
            
        Returns:
            Sheet name (string)
            
        Raises:
            ValueError: If sheet not found
        """
        try:
            spreadsheet = self.drive_service.sheets_service.spreadsheets().get(
                spreadsheetId=spreadsheet_id
            ).execute()
            
            for sheet in spreadsheet.get('sheets', []):
                if sheet['properties']['sheetId'] == sheet_id:
                    return sheet['properties']['title']
            
            raise ValueError(f"Sheet with ID {sheet_id} not found in spreadsheet {spreadsheet_id}")
        except Exception as e:
            print(f"⚠️  Error getting sheet name from ID: {e}", file=sys.stderr)
            raise
    
    def _read_column_from_google_sheet(
        self, 
        spreadsheet_id: str, 
        range_name: str, 
        sheet_id: int = None,
        skip_header_rows: int = 0,
        extract_column_index: int = None,
        flatten: bool = True
    ) -> list:
        """
        Generic method to read column data from Google Sheets.
        
        Args:
            spreadsheet_id: Google Sheets spreadsheet ID
            range_name: Range to read (e.g., "A:A", "A1:A100", "D2:D", or "'Sheet1'!A:A" if sheet name is included)
            sheet_id: Optional sheet ID (integer). If provided, will convert to sheet name for the range.
                      If None and range_name doesn't include sheet name, uses default sheet
            skip_header_rows: Number of header rows to skip (default 0)
            extract_column_index: If specified, extract only this column index (0-based). 
                                  If None, returns all columns as nested lists
            flatten: If True, flatten the result to a single list. Works with or without extract_column_index.
        
        Returns:
            List of values. If extract_column_index is set, returns a flat list of that column's values.
            If flatten is True and extract_column_index is None, returns a flat list of all values.
            Otherwise, returns a list of rows (each row is a list of cell values).
        """
        try:
            # Check if range_name already includes a sheet name (contains '!')
            if '!' in range_name:
                # Range already includes sheet name, use it as-is
                full_range = range_name
            elif sheet_id is not None:
                # Convert sheet_id to sheet_name and construct range
                sheet_name = self._get_sheet_name_from_id(spreadsheet_id, sheet_id)
                full_range = f"'{sheet_name}'!{range_name}"
            else:
                # Use range as-is (default sheet)
                full_range = range_name
            
            print(f"📖 Reading from Google Sheet {spreadsheet_id}, range: {full_range}", file=sys.stderr)
            
            result = self.drive_service.sheets_service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id,
                range=full_range
            ).execute()
            
            values = result.get('values', [])
            
            # Skip header rows if specified
            if skip_header_rows > 0 and len(values) > skip_header_rows:
                values = values[skip_header_rows:]
            
            # Extract specific column if requested
            if extract_column_index is not None:
                extracted = []
                for row in values:
                    if row and len(row) > extract_column_index:
                        extracted.append(row[extract_column_index])
                    else:
                        extracted.append('')  # Empty cell if row is too short
                return extracted
            
            # Flatten if requested
            if flatten:
                flattened = []
                for row in values:
                    if row:
                        flattened.extend(row)
                return flattened
            
            return values
            
        except Exception as e:
            print(f"⚠️  Error reading from Google Sheet: {e}", file=sys.stderr)
            raise
    
    def _get_allowed_gaps_list(self) -> set:
        """
        Read allowed gaps from Google Sheet and filter for 4-digit values.
        
        Returns:
            Set of allowed gap values (4-digit strings)
        """
        if not self.allowed_gaps_sheet_config:
            print("⚠️  Allowed gaps sheet config is not set, returning empty set", file=sys.stderr)
            return set()
        
        wb_id = self.allowed_gaps_sheet_config.get('wb_id')
        sheet_id = self.allowed_gaps_sheet_config.get('sheet_id')
        content_column_letter = self.allowed_gaps_sheet_config.get('content_column_letter', 'A')
        
        if not wb_id or sheet_id is None:
            print(f"⚠️  Allowed gaps sheet config missing wb_id or sheet_id: {self.allowed_gaps_sheet_config}", file=sys.stderr)
            return set()
        
        try:
            # Read column from row 1 onwards (we'll skip row 1 which is the header)
            column_range = f"{content_column_letter}1:{content_column_letter}"
            
            # Read the column, skip first row (header), and flatten
            all_values = self._read_column_from_google_sheet(
                spreadsheet_id=wb_id,
                range_name=column_range,
                sheet_id=sheet_id,
                skip_header_rows=1,  # Skip header row
                extract_column_index=0,  # Extract first (and only) column
            )
            
            # Filter for 4-digit values only
            filtered_values = self._filter_four_digit_cells(all_values)
            
            print(f"✅ Found {len(filtered_values)} allowed gaps (4-digit values)", file=sys.stderr)
            
            return set(filtered_values)
            
        except Exception as e:
            print(f"⚠️  Error reading allowed gaps sheet: {e}", file=sys.stderr)
            return set()
    
    def _filter_four_digit_cells(self, cell_values):
        """
        Filter and return only those cell values that contain exactly 4 digits (ignoring leading/trailing whitespace).
        Handles both string and numeric values from Google Sheets.
        
        Args:
            cell_values: Iterable of cell values (strings, numbers, or None)
            
        Returns:
            List of strings with exactly 4 digits.
        """
        filtered = []
        for v in cell_values:
            if v is None:
                continue
            
            # Convert to string and strip whitespace
            v_str = str(v).strip()
            
            # Check if it's exactly 4 digits (all characters are digits and length is 4)
            if v_str.isdigit() and len(v_str) == 4:
                filtered.append(v_str)
        return filtered

    def _get_customers(self, customers_input_file: Optional[Dict[str, Any]] = None):
        """
        Read column A from row 2 onwards from input Excel file or Google Sheets and return as a list.
        
        Args:
            customers_input_file: Either:
                - None: use latest customers file from Google Drive
                - Dict with keys:
                    - file_path: local path to an uploaded Excel file
                    - file_name: optional original file name (for display/logging)
            
        Returns:
            List of values from column A starting from row 2
        """
        if customers_input_file is None:
            customers_input_file_id, customers_input_file_name, first_sheet_id = self.get_last_customers_file()
            customers_input_file = customers_input_file_name
            customers = self._get_customers_from_google_drive(customers_input_file_id, first_sheet_id)
            return customers, customers_input_file

        if "file_path" not in customers_input_file or customers_input_file.get('file_path') is None:
            raise ValueError("customers_input_file['file_path'] is required when customers_input_file is provided")

        if "file_name" not in customers_input_file or customers_input_file.get('file_name') is None:
            raise ValueError("customers_input_file['file_name'] is required when customers_input_file is provided")
    

        customers_input_file_path = customers_input_file.get('file_path')
        customers_input_file_name = customers_input_file.get('file_name')

        # sys.path is a list (import paths) — use filesystem check instead
        if not os.path.exists(str(customers_input_file_path)):
            raise ValueError(f"File not found: {customers_input_file_path}")

        # Try reading as Excel file
        customers = self._get_customers_from_excel_file(customers_input_file_path)
        
        return customers, customers_input_file_name

    def _get_customers_from_excel_file(self, customers_input_file):
        """
        Read column A from row 2 onwards from an Excel file path.
        
        Args:
            customers_input_file: Path to Excel file (string)
            
        Returns:
            List of values from column A starting from row 2
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file is not a valid Excel file
        """
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
        
        print(f"Reading customers from Excel file: {customers_input_file}", file=sys.stderr)
        
        # Validate file exists
        if not os.path.exists(customers_input_file):
            raise FileNotFoundError(f"Excel file not found: {customers_input_file}")
        
        # Check file extension (allow temp files without extensions from PHP uploads)
        file_ext = os.path.splitext(customers_input_file)[1].lower()
        
        # Validate file extension if present
        if file_ext and file_ext not in ('.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'):
            raise ValueError(f"File is not a supported Excel format: {customers_input_file}. Supported: .xlsx, .xlsm, .xltx, .xltm, .xls")
        elif not file_ext:
            # No extension - likely a temp file from PHP upload, try to read it anyway
            print(f"File has no extension, attempting to read as Excel file", file=sys.stderr)
        
        try:
            # Try to open the file - openpyxl will validate the actual file format
            # This will raise InvalidFileException if it's not a valid Excel file
            input_wb = load_workbook(customers_input_file, read_only=True)
        except InvalidFileException as e:
            raise ValueError(f"Invalid Excel file format. Please ensure the file is a valid Excel file (.xlsx, .xlsm, .xltx, .xltm, .xls). Error: {e}")
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {e}")
        
        input_ws = input_wb.active
        
        # Read column A from row 2 onwards
        customers = []
        row = 2
        while True:
            cell_value = input_ws.cell(row=row, column=1).value
            if cell_value is None:
                break
            customers.append(cell_value)
            row += 1
        
        input_wb.close()
        return customers

    def _get_customers_from_google_sheets_id(self, file_id: str, sheet_id: int):
        """
        Get customers from Google Sheets using file ID.
        
        Args:
            file_id: Google Sheets file ID
            
        Returns:
            List of values from column A starting from row 2
        """
        # Get all data from column A (A:A gets the entire column)
        main_col_range = "A:A"
        return self._read_column_from_google_sheet(
            spreadsheet_id=file_id,
            range_name=main_col_range,
            sheet_id=sheet_id,
            skip_header_rows=1,  # Skip header row
            extract_column_index=0,  # Extract column A (first column)
        )
    
    def _get_customers_from_google_drive(self, customers_input_file_id, sheet_id: int = None):
        """
        Get customers from Google Drive using file ID.
        
        Args:
            customers_input_file_id: Google Sheets file ID

        Returns:
            List of values from column A starting from row 2
        """
        print(f"Customers input file: {customers_input_file_id}", file=sys.stderr)

        if customers_input_file_id is None:
            raise ValueError("No customers file found in Google Drive and no input file provided")

        # Get all data from column A (A:A gets the entire column)
        main_col_range = "A:A"
        return self._read_column_from_google_sheet(
            spreadsheet_id=customers_input_file_id,
            range_name=main_col_range,
            sheet_id=sheet_id,
            skip_header_rows=1,  # Skip header row
            extract_column_index=0,  # Extract column A (first column)
        )
    
    def get_last_customers_file(self):
        """
        Get the latest customers file ID from Google Drive folder that matches the filename pattern.
        
        Returns:
            File ID (str) of the latest matching file, or None if not found
            
        Raises:
            ValueError: If customers_google_folder_id is not set
        """
        if self.customers_google_folder_id is None:
            raise ValueError("Customers Google folder ID is not set")
        
        if self.customers_file_name_pattern is None:
            raise ValueError("Customers file name pattern is not set")

        print(f"Customers Google folder ID: {self.customers_google_folder_id}, Customers file name pattern: {self.customers_file_name_pattern}", file=sys.stderr)
        
        # Get the latest file matching the pattern from Google Drive
        result = self.drive_service.get_latest_file_by_pattern(
            folder_id=self.customers_google_folder_id,
            file_name_pattern=self.customers_file_name_pattern
        )
        
        if result is None:
            return None, None
        
        latest_file_id, latest_file_name, first_sheet_id = result
        return latest_file_id, latest_file_name, first_sheet_id

    def get_list_of_missing_customers(self, spread_sheet_id: str, sheet_id: int):
        """
        Get list of missing customers from the filter workbook.
        
        Args:
            spread_sheet_id: Google Sheets spreadsheet ID
            
        Returns:
            List of missing customer values (flattened)
        """
        if "filter" in self.excel_to_google_workbook:
            filter_workbook = self.excel_to_google_workbook["filter"]
            
            main_col_range = filter_workbook.get_summary_missing_customers_range()
            
            # Read the range and flatten the values
            return self._read_column_from_google_sheet(
                spreadsheet_id=spread_sheet_id,
                range_name=main_col_range,
                sheet_id=sheet_id,
                skip_header_rows=0,
                extract_column_index=None,  # Get all columns
            )

        else:
            print(f"FilterWorkbook not found in excel_to_google_workbook", file=sys.stderr)
            return None


def create_filter_google_manager(config_manager: ConfigManager):
    from .filter_file import FilterFile
    from .google_drive_utils import GDriveService
    config = _get_default_config(config_manager)
    print(f"Config created", file=sys.stderr)
    service_config = config.get_service_config()
    drive_service = GDriveService(service_config)

    module_name = 'filter'

    print(f"Drive service created", file=sys.stderr)

    output_files_config = config.get_output_files_config(module_name)

    spreadsheet_updaters = []

    for sheet_name_key, sheet_config in output_files_config.items():
        spreadsheet_updater = GapSpreadsheetUpdater(drive_service, sheet_config, sheet_name_key)
        spreadsheet_updaters.append(spreadsheet_updater)

    allowed_gaps_sheet_config = config.get_input_files_config(module_name).get('allowed_gaps_sheet', {})

    customers_google_folder_id = config.get_google_folder_id_by_name('customers', 'intermidiate')

    customers_file_name_pattern = config.get_output_file_pattern_by_name('customers', 'intermidiate')

    auto_dialer_file_name_pattern = config.get_excel_workbooks_config_by_name('customers').get('auto_dialer', {}).get('file_name_pattern', '')

    print(f"Customers Google folder ID: {customers_google_folder_id}, Customers file name pattern: {customers_file_name_pattern} auto dialer file name pattern: {auto_dialer_file_name_pattern}", file=sys.stderr)

    # Create mail service if mail config exists
    mail_config = config.get_mail_config_by_name(module_name)
    mail_service = create_mail_service(module_name, mail_config, service_config) if mail_config else None

    return FilterFile(drive_service, config_manager, module_name, spreadsheet_updaters, mail_service, customers_google_folder_id, customers_file_name_pattern, auto_dialer_file_name_pattern, allowed_gaps_sheet_config)