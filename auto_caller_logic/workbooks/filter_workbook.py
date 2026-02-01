"""
Filter workbook module.

Defines the Excel file structure for filter files.
"""

import io
from .base_workbook import ExcelToGoogleWorkbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import sys
import os
from datetime import datetime

class FilterWorkbook(ExcelToGoogleWorkbook):
    """Workbook for filter files."""

    def __init__(self, google_sheet_folder_id: str, excel_file_pattern: str, google_wb_name: str):
        super().__init__(google_sheet_folder_id, excel_file_pattern, google_wb_name)
        self._formulas = {}
        
        self.main_sheet_name = "פילטר חייגן"  # Store formulas to add after upload
        self.summary_sheet_name = "טיוטה"
    
    def create_excel_file(self, **kwargs):
        """
        Create Excel file for filter workbook.
        
        Args:
            data: List of data to write to Excel
            
        Returns:
            BytesIO buffer containing the Excel file
        """
        calls = kwargs.get('calls')
        customers = kwargs.get('customers')
        summarize_data = kwargs.get('summarize_data')
        
        
        if calls is None:
            raise ValueError("calls is required")
        if customers is None:
            raise ValueError("customers is required")
        if summarize_data is None:
            raise ValueError("summarize_data is required")

        try:

            # Create a new workbook
            wb = Workbook()
            ws = wb.active

            ws.title = self.main_sheet_name

            # Set RTL (Right-to-Left) direction
            ws.sheet_view.rightToLeft = True

            self._create_headers(ws)

            if calls is not None and len(calls) > 0:
                print(f"Storing {len(calls)} calls", file=sys.stderr)
                self._store_calls(ws, calls)
            if customers is not None and len(customers) > 0:
                print(f"Storing {len(customers)} customers", file=sys.stderr)
                self._store_customers(ws, customers)
            
            # Add current date and time in A1 with format: %d%m%y %H:%M
            ws_summary = wb.create_sheet(title=self.summary_sheet_name)

            ws_summary.sheet_view.rightToLeft = True

            # Use summarize_data passed from generate_data
            date_str = summarize_data.get('date_str', '')
            time_str = summarize_data.get('time_str', '')
            customers_input_file_val = summarize_data.get('customers_input_file_name', '')
            caller_id_val = summarize_data.get('caller_id', '')
            nick_name_val = summarize_data.get('nick_name', '')

            ws_summary['A1'] = date_str
            ws_summary['A2'] = time_str
            ws_summary['A3'] = customers_input_file_val
            ws_summary['A4'] = caller_id_val

            # Calculate the longest string among the values in column A
            values = [
                str(date_str) if date_str else '',
                str(time_str) if time_str else '',
                str(customers_input_file_val) if customers_input_file_val else '',
                str(caller_id_val) if caller_id_val else ''
            ]
            max_length = max(len(val) for val in values) if values else 10
            
            # Set column A width to fit the longest string (add 2 characters for padding)
            ws_summary.column_dimensions['A'].width = max_length + 2
            
            if f'{self.summary_sheet_name}!A6' not in self._formulas:
                self._formulas[f'{self.summary_sheet_name}!A6'] = f'=ARRAYFORMULA(SORT(UNIQUE(FILTER(\'{self.main_sheet_name}\'!H2:H, \'{self.main_sheet_name}\'!H2:H <> "")), 1, TRUE))'

            # Save to bytes buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            return excel_buffer
            
        except Exception as e:
            print(f"Error creating Filter Excel workbook: {e}", file=sys.stderr)
            raise RuntimeError(f"Error creating Filter Excel workbook: {e}")
    
    def _create_headers(self, ws):
        headers = [
        "שם מהחייגן",  # Column A: Name from Dialer
        "שם קבוצה",    # Column B: Group Name
        "שם יעד",       # Column C: Destination Name
        "שם קבוצה מפולטר",  # Column D: Filtered Group Name
        "שם יעד מפולטר",    # Column E: Filtered Destination Name
        "יש בחייגן אין בשם הקבוצה",  # Column F: Exists in Dialer, Not in Group Name
        "יש בחייגן אין בשם יעד",     # Column G: Exists in Dialer, Not in Destination Name
        "אין בשם קבוצה ואין בשם יעד"  # Column H: Not in Group Name and Not in Destination Name
         ]
        # Font style for headers: size 14, bold
        font_size = 14
        header_font = Font(size=font_size, bold=True)
        
        # Alignment style: wrap text
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        # Excel column width is measured in characters of default font (11pt)
        # Scale width calculation based on font size
        # Formula: base_width * (font_size / default_font_size) * hebrew_multiplier
        default_font_size = 11
        hebrew_multiplier = 1.2  # Hebrew characters are wider
        width_multiplier = (font_size / default_font_size) * hebrew_multiplier
        max_width = 30  # Maximum column width
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.alignment = wrap_alignment
            
            # Set column width based on text length and font size, with max limit
            column_width = min(max(len(header) * width_multiplier, 10), max_width)  # Min 10, Max 30
            ws.column_dimensions[cell.column_letter].width = column_width
        
        # Formulas will be added via Sheets API after upload
        # Store formulas with full ranges (sheet!cell) for later use
        sheet_name = "פילטר חייגן"
        self._formulas = {
            # Column D: Extract 4-digit numbers from column B
            # Note: Backslashes need to be escaped for Google Sheets API
            f'{sheet_name}!D2': '=ARRAYFORMULA(IFERROR(REGEXEXTRACT(B2:B & "", "(\\s?\\d{4})\\s?"), ""))',
            # Column E: Extract 4-digit numbers from column C
            f'{sheet_name}!E2': '=ARRAYFORMULA(IFERROR(REGEXEXTRACT(C2:C & "", "(\\s?\\d{4})\\s?"), ""))',
            # Column F: Exists in Dialer, Not in Group Name
            f'{sheet_name}!F2': '=ARRAYFORMULA(IF(A2:A="","",IF(H2:H = "", IF(COUNTIF(D:D,A2:A)> 0, "", TEXT(A2:A, "0")), "")))',
            # Column G: Exists in Dialer, Not in Destination Name
            f'{sheet_name}!G2': '=ARRAYFORMULA(IF(A2:A="","",IF(H2:H = "", IF(COUNTIF(E:E,A2:A)> 0, "", TEXT(A2:A, "0")), "")))',
            # Column H: Not in Group Name and Not in Destination Name
            f'{sheet_name}!H2': '=ARRAYFORMULA(IF(A2:A = "", "", IF((COUNTIF(E:E, A2:A) = 0) * (COUNTIF(D:D, A2:A) = 0), TEXT(A2:A, "0"), "")))'
        }

    def _store_calls(self, ws, calls):
        # Alignment style: don't wrap text, hide overflow (content that doesn't fit will be hidden)
        no_wrap_alignment = Alignment(wrap_text=False, shrink_to_fit=False)
        
        # Store the call 'name' in column B from row 2 and below
        for idx, call in enumerate(calls, start=2):
            cell = ws.cell(row=idx, column=2, value=call.get('NAME', ''))
            cell.alignment = no_wrap_alignment

    def _store_customers(self,ws, customers):
        # Alignment style: don't wrap text, hide overflow (content that doesn't fit will be hidden)
        no_wrap_alignment = Alignment(wrap_text=False, shrink_to_fit=False)
        
        # Store the customer 'name' in column A from row 2 and below
        for idx, customer in enumerate(customers, start=2):
            cell = ws.cell(row=idx, column=1, value=customer)
            cell.alignment = no_wrap_alignment

    def get_summary_missing_customers_range(self):
        return f"'{self.summary_sheet_name}'!A6:A"
    
    @staticmethod
    def get_header_values(customers_input_file: str, caller_id: str) -> list:
        """
        Get header values for A1-A4 (date, time, customers_input_file, caller_id).
        This is a static method to avoid code duplication.
        
        Args:
            customers_input_file: Name of the customers input file
            caller_id: ID of the caller
            
        Returns:
            List of 4 values: [date_str, time_str, customers_input_file, caller_id]
        """
        from datetime import datetime
        current_datetime = datetime.now()
        date_str = current_datetime.strftime("%d.%m.%Y")
        time_str = current_datetime.strftime("%H.%M")
        return [date_str, time_str, customers_input_file, caller_id]