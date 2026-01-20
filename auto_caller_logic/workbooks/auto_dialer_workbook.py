"""
Auto Dialer workbook module.

Defines the Excel file structure for auto dialer files.
"""

import io, sys
from .base_workbook import ExcelToGoogleWorkbook


class AutoDialerWorkbook(ExcelToGoogleWorkbook):
    """Workbook for auto dialer files."""
    
    def __init__(self, google_sheet_folder_id: str, excel_file_pattern: str, google_wb_name: str, output_folder_path: str):
        super().__init__(google_sheet_folder_id, excel_file_pattern, google_wb_name, output_folder_path)
        self._formulas = {}  # Store formulas to add after upload

    def create_excel_file(self, **kwargs):
        data = kwargs.get('customers')
        if data is None:
            raise ValueError("data is required")

        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            
            # Set sheet to RTL (Right-to-Left) direction
            ws.sheet_view.rightToLeft = True
            
            # Set headers in row 1
            headers = {
                'A1': 'מספרים בלי כוכבית',
                'B1': 'מספרים עם כוכבית',
                'C1': ''  # Empty title for column C
            }
            
            # Set header values and adjust column widths
            for cell_address, header_text in headers.items():
                ws[cell_address] = header_text
                if header_text:  # Only adjust width if there's text
                    # Calculate width based on text length (with multiplier for Hebrew characters)
                    column_letter = cell_address[0]
                    width = max(len(header_text) * 1.3, 10)  # Minimum width of 10
                    ws.column_dimensions[column_letter].width = width
            
            # Write data starting from row 2
            for idx, value in enumerate(data, start=2):
                # Column A: data value
                ws[f'A{idx}'] = value
                
                # Column B: Add asterisk before the value (literal string, not formula)
                ws[f'B{idx}'] = f'*{value}'
                
                # Column C: row number (line number)
                ws[f'C{idx}'] = idx - 1

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            return excel_buffer

            
        except ImportError:
            print(f"⚠️  Warning: openpyxl not available. Cannot create Excel file.", file=sys.stderr)
    
    