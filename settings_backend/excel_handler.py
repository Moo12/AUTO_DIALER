"""
Excel Handler for MySQL Integration

This module provides functionality to convert between Excel files and MySQL tables:
- Upload Excel rows to MySQL tables (with INSERT/UPDATE support)
- Export MySQL tables to Excel files
- Validate Excel structure against table schema
- Field value conversion (e.g., RGB to color name)
"""

import sys
import os
import re
import io
from pathlib import Path
from typing import Dict, Any, Optional, List, Callable, Union
import tempfile

try:
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError as e:
    print(f"⚠️  Required libraries not installed: {e}", file=sys.stderr)
    print("   Install with: pip install pandas openpyxl", file=sys.stderr)
    raise

from common_utils.db_connection import DatabaseConnection


class FieldConverterRegistry:
    """
    Registry for field value converters.
    
    Allows registering custom conversion functions that transform data values
    before displaying them in Excel.
    """
    
    # RGB tuple to color name mapping (used for both directions)
    RGB_TO_COLOR_NAME = {
        (0, 0, 0): 'Black',
        (255, 255, 255): 'White',
        (255, 0, 0): 'Red',
        (0, 255, 0): 'Green',
        (0, 0, 255): 'Blue',
        (255, 0, 255): 'Magenta',
        (0, 255, 255): 'Cyan',
        (192, 192, 192): 'Silver',
        (128, 0, 0): 'Maroon',
        (128, 128, 0): 'Olive',
        (0, 100, 0): 'Dark Green',  # #006400
        (128, 0, 128): 'Purple',
        (0, 128, 128): 'Teal',
        (0, 0, 128): 'Navy',
        (255, 165, 0): 'Orange',
        (50, 205, 50): 'Yellow',
        (169, 169, 169): 'Gray',
    }
    
    # Hebrew color name to English color name mapping
    HEBREW_TO_ENGLISH_COLOR = {
        'צהוב': 'Yellow',
        'אפור': 'Gray',
        'כתום': 'Orange',
        'אדום': 'Red',
        'ירוק': 'Green',
        'כחול': 'Blue',
        'סגול': 'Magenta',
        'כסף': 'Silver',
        'זהב': 'Olive',
        'ירוק כהה': 'Dark Green',
        'שחור': 'Black',
        'לבן': 'White',
    }
    
    @classmethod
    def _rgb_to_hex(cls, r: int, g: int, b: int) -> str:
        """
        Convert RGB values to hex format.
        
        Args:
            r, g, b: RGB values (0-255)
            
        Returns:
            Hex string in format "#RRGGBB"
        """
        return f"#{r:02x}{g:02x}{b:02x}".upper()
    
    @classmethod
    def _get_color_name_to_rgb(cls) -> Dict[str, str]:
        """
        Generate color name to hex RGB mapping from RGB_TO_COLOR_NAME constant.
        
        Returns:
            Dictionary mapping color names to hex RGB strings in format "#RRGGBB"
        """
        return {
            color_name: cls._rgb_to_hex(r, g, b)
            for (r, g, b), color_name in cls.RGB_TO_COLOR_NAME.items()
        }
    
    def __init__(self):
        """Initialize converter registry with built-in converters."""
        self._converters: Dict[str, Callable[[Any], Any]] = {}
        self._register_builtin_converters()
    
    def _register_builtin_converters(self):
        """Register built-in converter functions."""
        # RGB to color name converter (for MySQL to Excel)
        self.register('rgb_to_color', self._rgb_to_color_name)
        self.register('rgb_to_color_hebrew', self._rgb_to_color_hebrew)
        
        # Color name to RGB converter (for Excel to MySQL)
        self.register('color_to_rgb', self._color_name_to_rgb)
        
        # Hebrew color name to RGB hex converter (for Excel to MySQL)
        self.register('color_hebrew_name_to_rgb_hex', self._color_hebrew_name_to_rgb_hex)
        
        # Boolean to Yes/No
        self.register('bool_to_yesno', self._bool_to_yesno)
        
        # Boolean to Hebrew
        self.register('bool_to_hebrew', self._bool_to_hebrew)
        
        # Date format converter (ISO to readable)
        self.register('date_format', self._date_format)
        
        # Null to empty string
        self.register('null_to_empty', self._null_to_empty)
    
    def register(self, name: str, converter_func: Callable[[Any], Any]) -> None:
        """
        Register a converter function.
        
        Args:
            name: Unique name for the converter (e.g., 'rgb_to_color')
            converter_func: Function that takes a value and returns converted value
                           Should handle None/null values gracefully
        """
        self._converters[name] = converter_func
    
    def get(self, name: str) -> Optional[Callable[[Any], Any]]:
        """
        Get a converter function by name.
        
        Args:
            name: Converter name
            
        Returns:
            Converter function or None if not found
        """
        return self._converters.get(name)
    
    def list_converters(self) -> List[str]:
        """Get list of all registered converter names."""
        return list(self._converters.keys())
    
    # Built-in converter functions
    def _rgb_to_color_name(self, value: Any) -> str:
        """
        Convert RGB string to color name.
        
        Args:
            value: RGB string in format "rgb(r,g,b)" or "r,g,b" or hex "#RRGGBB"
            
        Returns:
            Color name or original value if conversion fails
        """
        if value is None or pd.isna(value):
            return ""
        
        value_str = str(value).strip()
        
        # Handle hex format (#RRGGBB)
        if value_str.startswith('#'):
            try:
                hex_color = value_str.lstrip('#')
                r = int(hex_color[0:2], 16)
                g = int(hex_color[2:4], 16)
                b = int(hex_color[4:6], 16)
                return self._rgb_to_name(r, g, b)
            except (ValueError, IndexError):
                return value_str
        
        # Handle rgb(r,g,b) format
        rgb_match = re.match(r'rgb\s*\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)', value_str, re.IGNORECASE)
        if rgb_match:
            r, g, b = map(int, rgb_match.groups())
            return self._rgb_to_name(r, g, b)
        
        # Handle comma-separated format (r,g,b)
        parts = value_str.split(',')
        if len(parts) == 3:
            try:
                r, g, b = map(int, [p.strip() for p in parts])
                return self._rgb_to_name(r, g, b)
            except ValueError:
                pass
        
        return value_str
    
    def _rgb_to_name(self, r: int, g: int, b: int) -> str:
        """
        Convert RGB values to color name.
        
        Args:
            r, g, b: RGB values (0-255)
            
        Returns:
            Color name
        """
        colors = self.RGB_TO_COLOR_NAME
        
        # Exact match
        if (r, g, b) in colors:
            return colors[(r, g, b)]
        
        # Find closest match by distance
        min_distance = float('inf')
        closest_color = None
        
        for (cr, cg, cb), name in colors.items():
            distance = ((r - cr) ** 2 + (g - cg) ** 2 + (b - cb) ** 2) ** 0.5
            if distance < min_distance:
                min_distance = distance
                closest_color = name
        
        # If very close, return the color name, otherwise return RGB string
        if min_distance < 30:  # Threshold for "close enough"
            return closest_color
        else:
            return f"RGB({r},{g},{b})"

    def _rgb_to_color_hebrew(self, value: Any) -> str:
        """
        Convert RGB string to color name in Hebrew.
        
        Args:
            value: RGB string in format "rgb(r,g,b)" or "r,g,b" or hex "#RRGGBB"
            
        Returns:
            Color name in Hebrew or original value if conversion fails
        """
        color_name = self._rgb_to_color_name(value)
        
        # Use the constant mapping (reverse lookup)
        # Create reverse mapping: English -> Hebrew
        english_to_hebrew = {v: k for k, v in self.HEBREW_TO_ENGLISH_COLOR.items()}
        
        # Return Hebrew name if found, otherwise return English name
        return english_to_hebrew.get(color_name, color_name)
    
    def _color_name_to_rgb(self, value: Any) -> str:
        """
        Convert color name to hex RGB string.
        
        This is the reverse of _rgb_to_color_name - converts color names
        (like "Red", "Blue", etc.) to hex RGB format "#RRGGBB".
        Also handles conversion from "r,g,b" format to hex.
        
        Args:
            value: Color name (e.g., "Red", "Blue", "Green", etc.) or RGB string "r,g,b"
            
        Returns:
            Hex RGB string in format "#RRGGBB" or original value if conversion fails
        """
        if value is None or pd.isna(value):
            return None
        
        value_str = str(value).strip()
        
        # If it's already in hex format, return as-is
        if value_str.startswith('#'):
            return value_str
        
        # If it's in "r,g,b" format, convert to hex
        rgb_match = re.match(r'^(\d+)\s*,\s*(\d+)\s*,\s*(\d+)$', value_str)
        if rgb_match:
            try:
                r, g, b = map(int, rgb_match.groups())
                return self._rgb_to_hex(r, g, b)
            except (ValueError, OverflowError):
                return value_str
        
        # If it's in rgb(r,g,b) format, convert to hex
        rgb_func_match = re.match(r'rgb\s*\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)', value_str, re.IGNORECASE)
        if rgb_func_match:
            try:
                r, g, b = map(int, rgb_func_match.groups())
                return self._rgb_to_hex(r, g, b)
            except (ValueError, OverflowError):
                return value_str
        
        # Get color name to hex RGB mapping from constant (reverse of RGB_TO_COLOR_NAME)
        color_to_hex = self._get_color_name_to_rgb()
        # Add alternative spelling for Gray/Grey
        color_to_hex['Grey'] = color_to_hex.get('Gray', self._rgb_to_hex(169, 169, 169))
        
        # Case-insensitive lookup
        color_name_lower = value_str.lower()
        for color_name, hex_value in color_to_hex.items():
            if color_name.lower() == color_name_lower:
                return hex_value
        
        # If not found, return original value
        return value_str
    
    def _color_hebrew_name_to_rgb_hex(self, value: Any) -> str:
        """
        Convert Hebrew color name to RGB hex string.
        
        Converts Hebrew color names (like "אדום", "כחול", etc.) to hex RGB format "#RRGGBB".
        
        Args:
            value: Hebrew color name (e.g., "אדום", "כחול", "ירוק", etc.)
            
        Returns:
            Hex RGB string in format "#RRGGBB" or original value if conversion fails
        """
        if value is None or pd.isna(value):
            return None
        
        value_str = str(value).strip()
        
        # If it's already in hex format, return as-is
        if value_str.startswith('#'):
            return value_str
        
        # If it's already in RGB format (r,g,b), convert to hex
        rgb_match = re.match(r'^(\d+)\s*,\s*(\d+)\s*,\s*(\d+)$', value_str)
        if rgb_match:
            try:
                r, g, b = map(int, rgb_match.groups())
                return self._rgb_to_hex(r, g, b)
            except (ValueError, OverflowError):
                return value_str
        
        # If it's in rgb(r,g,b) format, convert to hex
        rgb_func_match = re.match(r'rgb\s*\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)', value_str, re.IGNORECASE)
        if rgb_func_match:
            try:
                r, g, b = map(int, rgb_func_match.groups())
                return self._rgb_to_hex(r, g, b)
            except (ValueError, OverflowError):
                return value_str
        
        # Convert Hebrew color name to English color name using constant
        hebrew_lower = value_str.lower()
        english_color = None
        
        for hebrew_name, english_name in self.HEBREW_TO_ENGLISH_COLOR.items():
            if hebrew_name.lower() == hebrew_lower:
                english_color = english_name
                break
        
        # If Hebrew color found, convert English name to RGB hex
        if english_color:
            # Get color name to hex RGB mapping from constant
            color_to_hex = self._get_color_name_to_rgb()
            hex_value = color_to_hex.get(english_color)
            if hex_value:
                return hex_value
        
        # If not found, return original value
        return value_str
    
    def _bool_to_yesno(self, value: Any) -> str:
        """Convert boolean to Yes/No."""
        if value is None or pd.isna(value):
            return ""
        if isinstance(value, bool):
            return "Yes" if value else "No"
        if isinstance(value, (int, float)):
            return "Yes" if value else "No"
        value_str = str(value).lower()
        if value_str in ('true', '1', 'yes', 'y', 'on'):
            return "Yes"
        if value_str in ('false', '0', 'no', 'n', 'off'):
            return "No"
        return str(value)
    
    def _bool_to_hebrew(self, value: Any) -> str:
        """Convert boolean to Hebrew (כן/לא)."""
        if value is None or pd.isna(value):
            return ""
        if isinstance(value, bool):
            return "כן" if value else "לא"
        if isinstance(value, (int, float)):
            return "כן" if value else "לא"
        value_str = str(value).lower()
        if value_str in ('true', '1', 'yes', 'y', 'on'):
            return "כן"
        if value_str in ('false', '0', 'no', 'n', 'off'):
            return "לא"
        return str(value)
    
    def _date_format(self, value: Any) -> str:
        """Format date to readable string."""
        if value is None or pd.isna(value):
            return ""
        try:
            if isinstance(value, str):
                # Try to parse common date formats
                from datetime import datetime
                for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%m/%d/%Y']:
                    try:
                        dt = datetime.strptime(value, fmt)
                        return dt.strftime('%d/%m/%Y %H:%M')
                    except ValueError:
                        continue
            return str(value)
        except Exception:
            return str(value)
    
    def _null_to_empty(self, value: Any) -> str:
        """Convert None/null to empty string."""
        if value is None or pd.isna(value):
            return ""
        return str(value)


# Global converter registry instance
_converter_registry = FieldConverterRegistry()


def get_converter_registry() -> FieldConverterRegistry:
    """Get the global converter registry instance."""
    return _converter_registry


class ExcelHandler:
    """
    Handles Excel <-> MySQL conversions with validation and error handling.
    """
    
    def __init__(self, db_connection: DatabaseConnection, converter_registry: Optional[FieldConverterRegistry] = None):
        """
        Initialize Excel handler.
        
        Args:
            db_connection: DatabaseConnection instance for MySQL operations
            converter_registry: Optional FieldConverterRegistry instance (uses global registry if None)
        """
        self.db = db_connection
        self.converter_registry = converter_registry or get_converter_registry()
    
    def _execute_insert_and_get_pk(
        self,
        insert_query: str,
        insert_params: Dict[str, Any],
        primary_key: Optional[str],
        pk_value: Any,
        is_auto_increment: bool
    ) -> Any:
        """
        Execute an INSERT operation and return the primary key value.
        
        Args:
            insert_query: The INSERT query string
            insert_params: Parameters for the INSERT query
            primary_key: Name of the primary key column (or None)
            pk_value: Primary key value from the row data (or None)
            is_auto_increment: Whether the primary key is auto_increment
            
        Returns:
            The primary key value of the inserted row
        """
        if not primary_key:
            # No primary key, just execute the insert
            self.db.execute_update(insert_query, insert_params)
            return None
        
        if is_auto_increment:
            # For auto_increment, execute insert and get last insert ID
            try:
                if not self.db._is_connected:
                    self.db.connect()
                with self.db.get_connection() as conn:
                    from sqlalchemy import text
                    result = conn.execute(text(insert_query), insert_params or {})
                    conn.commit()
                    # Get the last insert ID
                    last_id = result.lastrowid
                    return last_id
            except Exception as e:
                print(f"Warning: Could not get last insert ID: {e}", file=sys.stderr)
                # Fallback: execute normally
                self.db.execute_update(insert_query, insert_params)
                return None
        else:
            # For non-auto_increment, use the provided primary key value
            self.db.execute_update(insert_query, insert_params)
            return pk_value
    
    def validate_excel_structure(
        self,
        file_path: str,
        table_name: str,
        sheet_name: Optional[str] = None,
        header_row: int = 0,
        start_row: int = 1,
        mapping: Optional[Dict[str, str]] = None
    ) -> Dict[str, Any]:
        """
        Validate Excel file structure against table schema.
        
        Args:
            file_path: Path to Excel file
            table_name: Name of MySQL table
            sheet_name: Name of sheet to read (None = first sheet)
            header_row: Row number containing headers (0-indexed)
            start_row: Row number where data starts (0-indexed, after header)
            mapping: Optional dictionary mapping Excel column names to MySQL column names.
                     Format: {Excel Column Name: MySQL Column Name}
                     Example: {"Name": "name", "Phone Number": "phone_number"}
                     Note: Column names are automatically stripped of leading/trailing whitespace.
            
        Returns:
            Dictionary with validation results:
                {
                    'valid': bool,
                    'errors': List[str],
                    'warnings': List[str],
                    'excel_columns': List[str],  # Original Excel column names
                    'mapped_columns': List[str],  # Column names after mapping (if mapping provided)
                    'table_columns': List[str],
                    'missing_columns': List[str],
                    'extra_columns': List[str]
                }
        """
        errors = []
        warnings = []
        
        # Validate file exists
        if not os.path.exists(file_path):
            return {
                'valid': False,
                'errors': [f"File not found: {file_path}"],
                'warnings': [],
                'excel_columns': [],
                'mapped_columns': [],
                'table_columns': [],
                'missing_columns': [],
                'extra_columns': []
            }
        
        # Get table schema
        try:
            table_schema = self.db.get_table_schema(table_name)
            if not table_schema:
                errors.append(f"Table '{table_name}' not found in database")
                return {
                    'valid': False,
                    'errors': errors,
                    'warnings': warnings,
                    'excel_columns': [],
                    'mapped_columns': [],
                    'table_columns': [],
                    'missing_columns': [],
                    'extra_columns': []
                }
        except Exception as e:
            errors.append(f"Error getting table schema: {str(e)}")
            return {
                'valid': False,
                'errors': errors,
                'warnings': warnings,
                'excel_columns': [],
                'mapped_columns': [],
                'table_columns': [],
                'missing_columns': [],
                'extra_columns': []
            }
        
        # Read Excel file
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header_row,
                nrows=0  # Only read headers
            )
            
            if isinstance(df, dict):
                # Multiple sheets, use first one
                df = list(df.values())[0]
            
            excel_columns = [str(col).strip() for col in df.columns.tolist()]
            table_columns = [col['column_name'] for col in table_schema]
            
            # Apply mapping if provided to get mapped column names
            mapped_columns = excel_columns.copy()
            if mapping:
                # Apply mapping: Excel column name -> MySQL column name
                mapped_columns = [mapping.get(col, col) for col in excel_columns]
                # Check for unmapped Excel columns that don't exist in table
                unmapped_excel_cols = [col for col in excel_columns if col not in mapping and col not in table_columns]
                if unmapped_excel_cols:
                    warnings.append(f"Excel columns not in mapping and not in table (will be ignored): {', '.join(unmapped_excel_cols)}")
            
            # Find primary key column
            primary_key = None
            for col in table_schema:
                if col['column_key'] == 'PRI':
                    primary_key = col['column_name']
                    break
            
            # Check for missing required columns (non-nullable, no default, not auto_increment)
            # Use mapped_columns for comparison
            missing_columns = []
            required_columns = []
            for col in table_schema:
                # Skip auto_increment columns (they are auto-generated)
                is_auto_increment = col.get('extra', '').lower() == 'auto_increment'
                if col['is_nullable'] == 'NO' and col['column_default'] is None and not is_auto_increment:
                    required_columns.append(col['column_name'])
            
            for req_col in required_columns:
                if req_col not in mapped_columns:
                    missing_columns.append(req_col)
                    # Find the original Excel column name if it was mapped
                    excel_col_name = None
                    if mapping:
                        for excel_col, mysql_col in mapping.items():
                            if mysql_col == req_col:
                                excel_col_name = excel_col
                                break
                    if excel_col_name:
                        errors.append(f"Missing required column: '{req_col}' (Excel column '{excel_col_name}' should map to this)")
                    else:
                        errors.append(f"Missing required column: '{req_col}'")
            
            # Check for extra columns (warnings only)
            # Use mapped_columns for comparison
            extra_columns = [col for col in mapped_columns if col not in table_columns]
            if extra_columns:
                warnings.append(f"Extra columns in Excel (will be ignored): {', '.join(extra_columns)}")
            
            # Check if primary key is present (for UPDATE operations)
            # Use mapped_columns for comparison
            if primary_key and primary_key not in mapped_columns:
                warnings.append(f"Primary key column '{primary_key}' not found. All rows will be inserted (no updates).")
            
            valid = len(errors) == 0
            
            return {
                'valid': valid,
                'errors': errors,
                'warnings': warnings,
                'excel_columns': excel_columns,
                'mapped_columns': mapped_columns,
                'table_columns': table_columns,
                'missing_columns': missing_columns,
                'extra_columns': extra_columns,
                'primary_key': primary_key
            }
            
        except InvalidFileException:
            errors.append(f"Invalid Excel file format: {file_path}")
            return {
                'valid': False,
                'errors': errors,
                'warnings': warnings,
                'excel_columns': [],
                'mapped_columns': [],
                'table_columns': [],
                'missing_columns': [],
                'extra_columns': []
            }
        except Exception as e:
            errors.append(f"Error reading Excel file: {str(e)}")
            return {
                'valid': False,
                'errors': errors,
                'warnings': warnings,
                'excel_columns': [],
                'mapped_columns': [],
                'table_columns': [],
                'missing_columns': [],
                'extra_columns': []
            }
    
    def excel_to_mysql(
        self,
        file_path: str,
        table_name: str,
        sheet_name: Optional[str] = None,
        mapping: Optional[Dict[str, str]] = None,
        header_row: int = 0,
        start_row: int = 1,
        update_on_duplicate: bool = True,
        batch_size: int = 100,
        column_converters: Optional[Dict[str, str]] = None
    ) -> Dict[str, Any]:
        """
        Upload Excel rows to MySQL table.
        
        Supports INSERT and UPDATE operations based on primary key.
        If primary key exists in Excel and update_on_duplicate is True, performs UPDATE.
        Otherwise, performs INSERT.
        
        Args:
            file_path: Path to Excel file
            table_name: Name of MySQL table
            sheet_name: Name of sheet to read (None = first sheet)
            mapping: Optional dictionary mapping Excel column names to MySQL column names.
                     Format: {Excel Column Name: MySQL Column Name}
                     Example: {"Name": "name", "Phone Number": "phone_number", "Email Title": "email_title"}
                     Note: Column names are automatically stripped of leading/trailing whitespace.
            header_row: Row number containing headers (0-indexed)
            start_row: Row number where data starts (0-indexed, after header)
            update_on_duplicate: If True, update existing rows based on primary key
            batch_size: Number of rows to process in each batch
            column_converters: Optional dictionary mapping MySQL column names to converter names
                             e.g., {'color_rgb': 'color_to_rgb', 'is_active': 'bool_to_yesno'}
                             Available converters: color_to_rgb, bool_to_yesno, bool_to_hebrew, date_format, null_to_empty
            
        Returns:
            Dictionary with operation results:
                {
                    'success': bool,
                    'rows_inserted': int,
                    'rows_updated': int,
                    'rows_skipped': int,
                    'errors': List[Dict[str, Any]],  # List of {row: int, error: str}
                    'total_rows': int
                }
        """
        errors = []
        rows_inserted = []  # List of primary keys for inserted rows
        rows_updated = []   # List of primary keys for updated rows
        rows_skipped = 0
        
        # Validate structure first
        validation = self.validate_excel_structure(
            file_path, table_name, sheet_name, header_row, start_row, mapping
        )
        
        if not validation['valid']:
            return {
                'success': False,
                'rows_inserted': 0,
                'rows_updated': 0,
                'rows_skipped': 0,
                'errors': [{'row': 0, 'error': err} for err in validation['errors']],
                'total_rows': 0
            }
        
        # Get table schema
        table_schema = self.db.get_table_schema(table_name)
        primary_key = validation.get('primary_key')
        
        # Check if primary key is auto_increment
        is_auto_increment = False
        if primary_key:
            for col in table_schema:
                if col['column_name'] == primary_key:
                    is_auto_increment = col.get('extra', '').lower() == 'auto_increment'
                    break

        print(f"header_row: {header_row}")
        print(f"start_row: {start_row}")
        
        # Read Excel file
        try:
            # Calculate skiprows: skip rows between header and start_row
            # Never skip the header row itself
            skiprows = None
            if start_row > header_row + 1:
                # Skip rows between header_row+1 and start_row-1 (inclusive)
                skiprows = range(header_row + 1, start_row)
            
            print(f"skiprows: {skiprows}")
            
            # Read Excel file with dtype=str to preserve leading zeros (e.g., phone numbers like "0502574817")
            # This prevents pandas from converting numeric-looking strings to numbers
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header_row,
                skiprows=skiprows,
                dtype=str  # Read all columns as strings to preserve leading zeros
            )
            
            if isinstance(df, dict):
                df = list(df.values())[0]
            
            if df.empty:
                return {
                    'success': True,
                    'rows_inserted': 0,
                    'rows_updated': 0,
                    'rows_skipped': 0,
                    'errors': [],
                    'total_rows': 0
                }
            
            # Strip whitespace from column names (to match validation behavior)
            df.columns = [str(col).strip() for col in df.columns]
            
            # Apply column mapping if provided
            # Mapping convention: {Excel Column Name: MySQL Column Name}
            # Example: {"Name": "name", "Phone Number": "phone_number"}
            if mapping:
                # Create a mapping with stripped keys to handle any whitespace issues
                stripped_mapping = {str(k).strip(): v for k, v in mapping.items()}
                df = df.rename(columns=stripped_mapping)

                print(f"stripped_mapping: {stripped_mapping}")
                print(f"df.columns: {df.columns}")
            
            # Filter to only include columns that exist in table
            table_columns = [col['column_name'] for col in table_schema]
            df = df[[col for col in df.columns if col in table_columns]]
            
            # Apply column converters if provided (before type conversion)
            if column_converters:
                # Validate that all converter columns exist in the DataFrame
                missing_cols = [col for col in column_converters.keys() if col not in df.columns]
                if missing_cols:
                    return {
                        'success': False,
                        'rows_inserted': [],
                        'rows_updated': [],
                        'rows_skipped': 0,
                        'errors': [{'row': 0, 'error': f"Cannot convert columns that don't exist in Excel: {', '.join(missing_cols)}"}],
                        'total_rows': 0
                    }
                
                # Validate converter names
                available_converters = self.converter_registry.list_converters()
                invalid_converters = [
                    f"{col}:{conv}" for col, conv in column_converters.items()
                    if conv not in available_converters
                ]
                if invalid_converters:
                    return {
                        'success': False,
                        'rows_inserted': [],
                        'rows_updated': [],
                        'rows_skipped': 0,
                        'errors': [{'row': 0, 'error': f"Invalid converter names: {', '.join(invalid_converters)}. Available converters: {', '.join(available_converters)}"}],
                        'total_rows': 0
                    }
                
                # Apply converters to each column
                for col_name, converter_name in column_converters.items():
                    converter_func = self.converter_registry.get(converter_name)
                    if converter_func:
                        df[col_name] = df[col_name].apply(converter_func)
            
            # Convert pandas "nan" strings (from reading as dtype=str) and empty strings to None
            # This handles the case where pandas converts NaN to the string "nan" when dtype=str
            df = df.replace(['nan', 'NaN', 'None', ''], None)
            
            # Convert numeric columns to appropriate types based on database schema
            # This preserves leading zeros for VARCHAR/TEXT columns (kept as strings)
            for col_info in table_schema:
                col_name = col_info['column_name']
                if col_name in df.columns:
                    data_type = col_info['data_type'].upper()
                    # Only convert to numeric if column is numeric type in database
                    # Keep VARCHAR/TEXT/CHAR as strings to preserve leading zeros (e.g., phone numbers)
                    if data_type in ('INT', 'INTEGER', 'BIGINT', 'SMALLINT', 'TINYINT', 'MEDIUMINT'):
                        # Convert to int, but preserve None values and handle conversion errors
                        def convert_to_int(x):
                            if x is None or str(x).strip() in ('', 'nan', 'NaN', 'None'):
                                return None
                            try:
                                return int(float(str(x)))
                            except (ValueError, TypeError):
                                return None
                        df[col_name] = df[col_name].apply(convert_to_int)
                    elif data_type in ('FLOAT', 'DOUBLE', 'DECIMAL', 'NUMERIC'):
                        # Convert to float, but preserve None values and handle conversion errors
                        def convert_to_float(x):
                            if x is None or str(x).strip() in ('', 'nan', 'NaN', 'None'):
                                return None
                            try:
                                return float(str(x))
                            except (ValueError, TypeError):
                                return None
                        df[col_name] = df[col_name].apply(convert_to_float)
                    # For VARCHAR, TEXT, CHAR, etc. - keep as string to preserve leading zeros
            
            # Final pass: replace any remaining NaN values with None
            df = df.where(pd.notnull(df), None)
            
            total_rows = len(df)
            
            # Process in batches
            for batch_start in range(0, total_rows, batch_size):
                batch_end = min(batch_start + batch_size, total_rows)
                batch_df = df.iloc[batch_start:batch_end]
                
                for idx, row in batch_df.iterrows():
                    row_num = batch_start + idx + start_row + 1  # Excel row number (1-indexed)
                    row_data = row.to_dict()
                    
                    try:
                        # Handle NULL values for columns that don't allow NULL
                        # If column has a default value, remove it from row_data so MySQL uses the default
                        # If column has no default value, skip the row with an error
                        skip_row = False
                        columns_to_remove = []  # Collect columns to remove (those with defaults)
                        
                        for col in table_schema:
                            col_name = col['column_name']
                            if col['is_nullable'] == 'NO' and col_name in row_data and row_data[col_name] is None:
                                column_default = col.get('column_default')
                                if column_default is None:
                                    # No default value - this is an error, skip the row
                                    errors.append({
                                        'row': row_num,
                                        'error': f"Column '{col_name}' cannot be NULL and has no default value"
                                    })
                                    skip_row = True
                                    break
                                else:
                                    # Column has a default value - mark it for removal from row_data
                                    # This way the column won't be included in the INSERT statement
                                    columns_to_remove.append(col_name)
                        
                        if skip_row:
                            rows_skipped += 1
                            continue
                        
                        # Remove columns with defaults from row_data so MySQL will use the defaults
                        # This ensures these columns are not included in the INSERT statement
                        for col_name in columns_to_remove:
                            if col_name in row_data:
                                del row_data[col_name]
                        
                        # Check if we should update or insert
                        if update_on_duplicate and primary_key and primary_key in row_data:
                            # Try to update first
                            pk_value = row_data[primary_key]
                            
                            # Build UPDATE query
                            update_cols = [col for col in row_data.keys() 
                                         if col != primary_key and col in table_columns]
                            
                            if update_cols:
                                set_clause = ', '.join([f"{col} = :{col}" for col in update_cols])
                                update_query = f"UPDATE {table_name} SET {set_clause} WHERE {primary_key} = :pk_value"
                                
                                update_params = {col: row_data[col] for col in update_cols}
                                update_params['pk_value'] = pk_value
                                
                                affected = self.db.execute_update(update_query, update_params)
                                
                                if affected > 0:
                                    rows_updated.append(pk_value)
                                    continue
                            
                            # If update didn't affect any rows, insert
                            insert_cols = [col for col in row_data.keys() if col in table_columns]
                            insert_query = f"INSERT INTO {table_name} ({', '.join(insert_cols)}) VALUES ({', '.join([f':{col}' for col in insert_cols])})"
                            insert_params = {col: row_data[col] for col in insert_cols}
                            
                            # Execute insert and get the inserted primary key value
                            inserted_pk = self._execute_insert_and_get_pk(
                                insert_query, insert_params, primary_key, pk_value, is_auto_increment
                            )
                            rows_inserted.append(inserted_pk)
                        else:
                            # Insert only
                            insert_cols = [col for col in row_data.keys() if col in table_columns]
                            print(f"raw row data: {row_data}")
                            insert_query = f"INSERT INTO {table_name} ({', '.join(insert_cols)}) VALUES ({', '.join([f':{col}' for col in insert_cols])})"
                            insert_params = {col: row_data[col] for col in insert_cols}
                            
                            # Execute insert and get the inserted primary key value
                            pk_value_from_row = row_data.get(primary_key) if primary_key else None
                            inserted_pk = self._execute_insert_and_get_pk(
                                insert_query, insert_params, primary_key, pk_value_from_row, is_auto_increment
                            )
                            rows_inserted.append(inserted_pk)
                            
                    except Exception as e:
                        error_msg = str(e)
                        errors.append({
                            'row': row_num,
                            'error': error_msg
                        })
                        rows_skipped += 1
                        print(f"Error processing row {row_num}: {error_msg}", file=sys.stderr)
            
            success = len(errors) == 0 or (len(rows_inserted) + len(rows_updated)) > 0

            print(f"errors: {errors}")
            print(f"rows_inserted: {rows_inserted}")
            print(f"rows_updated: {rows_updated}")
            print(f"rows_skipped: {rows_skipped}")
            print(f"total_rows: {total_rows}")
            print(f"success: {success}")
            
            return {
                'success': success,
                'rows_inserted': rows_inserted,
                'rows_updated': rows_updated,
                'rows_skipped': rows_skipped,
                'errors': errors,
                'total_rows': total_rows
            }
            
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}", file=sys.stderr)
            return {
                'success': False,
                'rows_inserted': 0,
                'rows_updated': 0,
                'rows_skipped': 0,
                'errors': [{'row': 0, 'error': f"Error reading Excel file: {str(e)}"}],
                'total_rows': 0
            }
    
    def mysql_to_excel(
        self,
        table_name: str,
        output_path: Optional[str] = None,
        filters: Optional[Dict[str, Any]] = None,
        columns: Optional[List[str]] = None,
        sheet_name: str = "Sheet1",
        column_mapping: Optional[Dict[str, str]] = None,
        column_converters: Optional[Dict[str, str]] = None
    ) -> Union[bytes, str]:
        """
        Export MySQL table to Excel file.
        
        Args:
            table_name: Name of MySQL table to export
            output_path: Optional path to save Excel file. If None, returns bytes content only
            filters: Optional dictionary of filters {column: value} for WHERE clause
            columns: Optional list of column names to export (if None, exports all)
            sheet_name: Name of Excel sheet
            column_mapping: Optional dictionary mapping database column names to display names
                          e.g., {'db_column_name': 'Display Name', 'user_id': 'User ID'}
            column_converters: Optional dictionary mapping column names to converter names
                             e.g., {'color_rgb': 'rgb_to_color', 'is_active': 'bool_to_yesno'}
                             Available converters: rgb_to_color, bool_to_yesno, bool_to_hebrew,
                             date_format, null_to_empty, or any custom registered converter
            
        Returns:
            If output_path is provided: Returns the output_path (str)
            If output_path is None: Returns Excel file content as bytes
            
        Raises:
            ValueError: If table doesn't exist or query fails
        """
        # Build SELECT query
        if columns:
            # Validate columns exist
            table_schema = self.db.get_table_schema(table_name)
            table_columns = [col['column_name'] for col in table_schema]
            invalid_cols = [col for col in columns if col not in table_columns]
            if invalid_cols:
                raise ValueError(f"Invalid columns: {', '.join(invalid_cols)}")
            select_cols = ', '.join(columns)
        else:
            select_cols = '*'
        
        query = f"SELECT {select_cols} FROM {table_name}"
        params = {}
        
        # Add WHERE clause if filters provided
        if filters:
            where_conditions = []
            for col, value in filters.items():
                where_conditions.append(f"{col} = :{col}")
                params[col] = value
            query += " WHERE " + " AND ".join(where_conditions)
        
        # Execute query
        try:
            rows = self.db.execute_query(query, params)
            
            if not rows:
                # Create empty Excel with headers
                df = pd.DataFrame(columns=columns or [])
            else:
                df = pd.DataFrame(rows)
            
            # Apply column converters if provided (before column mapping)
            if column_converters:
                # Validate that all converter columns exist in the DataFrame
                missing_cols = [col for col in column_converters.keys() if col not in df.columns]
                if missing_cols:
                    raise ValueError(
                        f"Cannot convert columns that don't exist in result: {', '.join(missing_cols)}"
                    )
                
                # Validate converter names
                available_converters = self.converter_registry.list_converters()
                invalid_converters = [
                    f"{col}:{conv}" for col, conv in column_converters.items()
                    if conv not in available_converters
                ]
                if invalid_converters:
                    raise ValueError(
                        f"Invalid converter names: {', '.join(invalid_converters)}. "
                        f"Available converters: {', '.join(available_converters)}"
                    )
                
                # Apply converters to each column
                for col_name, converter_name in column_converters.items():
                    converter_func = self.converter_registry.get(converter_name)
                    if converter_func:
                        df[col_name] = df[col_name].apply(converter_func)
            
            # Apply column name mapping if provided (after conversion)
            if column_mapping:
                # Validate that all mapped columns exist in the DataFrame
                missing_cols = [col for col in column_mapping.keys() if col not in df.columns]
                if missing_cols:
                    raise ValueError(
                        f"Cannot map columns that don't exist in result: {', '.join(missing_cols)}"
                    )
                # Rename columns using the mapping
                df = df.rename(columns=column_mapping)
            
            # Create Excel file in memory
            excel_buffer = io.BytesIO()
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the workbook and worksheet to format timestamp columns
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Get table schema to identify timestamp/datetime columns
                table_schema = self.db.get_table_schema(table_name)
                timestamp_columns = []
                
                for col_info in table_schema:
                    col_name = col_info['column_name']
                    data_type = col_info.get('data_type', '').upper()
                    
                    # Check if column is TIMESTAMP, DATETIME, or DATE type
                    if data_type in ('TIMESTAMP', 'DATETIME', 'DATE'):
                        # Find the column index in the DataFrame (after mapping if applied)
                        if col_name in df.columns:
                            timestamp_columns.append(col_name)
                        elif column_mapping:
                            # Check if column was renamed
                            for orig_col, mapped_col in column_mapping.items():
                                if orig_col == col_name and mapped_col in df.columns:
                                    timestamp_columns.append(mapped_col)
                                    break
                
                # Format timestamp columns and adjust column widths
                from openpyxl.styles import numbers
                from datetime import datetime
                
                for col_idx, col_name in enumerate(df.columns, start=1):
                    if col_name in timestamp_columns:
                        # Set datetime format for timestamp columns
                        # Format: "yyyy-mm-dd hh:mm:ss" or "yyyy-mm-dd" for DATE type
                        for col_info in table_schema:
                            if col_info['column_name'] == col_name or \
                               (column_mapping and column_mapping.get(col_info['column_name']) == col_name):
                                data_type = col_info.get('data_type', '').upper()
                                if data_type == 'DATE':
                                    date_format = 'yyyy-mm-dd'
                                else:
                                    date_format = 'yyyy-mm-dd hh:mm:ss'
                                break
                        else:
                            date_format = 'yyyy-mm-dd hh:mm:ss'
                        
                        # Apply format to all cells in the column (skip header row)
                        for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if cell.value is not None:
                                # Ensure the cell is formatted as datetime
                                cell.number_format = date_format
                        
                        # Set column width to accommodate datetime format
                        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                        # Get current width if column dimension exists, otherwise use default
                        current_width = 10
                        if col_letter in worksheet.column_dimensions:
                            col_dim = worksheet.column_dimensions[col_letter]
                            if col_dim.width:
                                current_width = col_dim.width
                        worksheet.column_dimensions[col_letter].width = max(20, current_width)
            
            # Get the bytes content after formatting (writer context manager saves automatically)
            excel_content = excel_buffer.getvalue()
            excel_buffer.close()
            
            # If output_path is provided, save to file
            if output_path is not None:
                # Ensure output directory exists
                output_dir = os.path.dirname(output_path)
                if output_dir:
                    os.makedirs(output_dir, exist_ok=True)
                
                # Write to file
                with open(output_path, 'wb') as f:
                    f.write(excel_content)
                
                print(f"✓ Exported {len(rows)} rows from '{table_name}' to {output_path}", file=sys.stderr)
                return output_path
            else:
                # Return bytes content
                print(f"✓ Exported {len(rows)} rows from '{table_name}' (in-memory)", file=sys.stderr)
                return excel_content
            
        except Exception as e:
            error_msg = f"Error exporting table '{table_name}' to Excel: {str(e)}"
            print(f"✗ {error_msg}", file=sys.stderr)
            raise ValueError(error_msg) from e

